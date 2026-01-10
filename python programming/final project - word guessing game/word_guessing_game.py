"""
word_guessing_game.py

Word guessing game that fetches words from a website and stores them in an Excel file.

Author: Christopher Romo & Alyssa Walker
Created: 2023-07-17
"""


import openpyxl
import random
import requests
import bs4


def guess_word() -> None:
    """Function to play the word guessing game."""

    # opens the excel workbook
    wb = openpyxl.load_workbook('random_words.xlsx')
    the_sheet = wb.active

    # gets the number of words in the sheet
    num_words = the_sheet.max_row

    # randomly picks a word
    random_row = random.randint(1, num_words)
    word = the_sheet.cell(row=random_row, column=1).value.strip()

    # hides the word with underscores
    display = ['_'] * len(word)
    print(' '.join(display))

    # keeps track of guess count
    guess_count = 0
    mistakes_count = 0

    # guess loop
    while guess_count < len(word) and mistakes_count < 5:
        guess = input("Enter your guess: ").lower()

        # ensures only a single character is entered
        if len(guess) != 1:
            print("Please enter a single letter!")
            continue

        found = False

        # displays guess if it was correct
        for i in range(len(word)):
            if word[i] == guess:
                display[i] = guess
                found = True
                guess_count += 1

        print(' '.join(display))

        # mistake display
        if not found:
            mistakes_count += 1
            print("Incorrect guess. Mistakes remaining:", 5 - mistakes_count)

    # check if the word has been guessed
    if '_' not in display:
        print("You Won!")
    else:
        print("You Lost! The word was:", word)

    # close the workbook
    wb.close()


def main() -> None:
    """Program entry point."""
    
    # opens the excel workbook
    wb = openpyxl.Workbook()
    the_sheet = wb.active

    # get website text and check for status
    web_one = requests.get('https://www.ef.edu/english-resources/english-vocabulary/top-1000-words/')
    web_one.raise_for_status()

    # parse website into recognizable html text
    web_one_par = bs4.BeautifulSoup(web_one.text, 'html.parser')

    # extract content from parsed text
    elements_one = str(web_one_par.find(class_='content'))

    # split the words into a list
    the_words = elements_one.split('<br/>')

    # fix beginning and last words via slicing
    the_words[0] = the_words[0][-1]
    the_words[999] = the_words[999][0:10]

    # counter variable for cell saving
    counter = 1

    # for loop adds each word into the next cell of excel sheet
    for element in the_words:
        the_cell = 'A' + str(counter)
        the_sheet[the_cell] = element
        counter += 1

    # saves the excel sheet
    wb.save('random_words.xlsx')

    # calls the guess_word function
    guess_word()


if __name__ == "__main__":
    main()