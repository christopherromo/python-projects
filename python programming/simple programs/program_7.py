"""
program_7.py

Openpyxl example.

Author: Christopher Romo
"""


import openpyxl


def main() -> None:
    """Program entry point."""

    # load an existing workbook and read values from it
    wb = openpyxl.load_workbook('test.xlsx')
    the_sheet = wb['Sheet1']
    items = the_sheet['A1':'A5']

    # display the cell coordinates and values
    for the_row in items:
        for the_cell in the_row:
            print(f'{the_cell.coordinate}: {the_cell.value}')

    # create a new workbook and write values to it
    wb = openpyxl.Workbook()

    # get the active sheet
    sheet = wb.active
    print(sheet.title)

    # rename the sheet
    sheet.title = 'Number and its square'
    print(wb.sheetnames)

    # write values to the sheet
    sheet['A1'] = 'Value'
    sheet['B1'] = 'Squared Value'

    cell_index = 2

    # write the values and their squares
    for the_row in items:
        col_a_name = 'A' + str(cell_index)
        sheet[col_a_name] = the_row[0].value

        col_b_name = 'B' + str(cell_index)
        sheet[col_b_name] = the_row[0].value ** 2

        cell_index += 1

    # save the new workbook
    wb.save('test_two.xlsx')


if __name__ == "__main__":
    main()